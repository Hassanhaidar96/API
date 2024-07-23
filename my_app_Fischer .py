# -*- coding: utf-8 -*-
"""
Created on Mon Jul 22 08:49:23 2024

@author: hah
"""
import streamlit as st
from sympy import symbols, Eq, solve
import math
from scipy.optimize import fsolve
from PIL import Image
from scipy.optimize import newton
import streamlit as st
from reportlab.pdfgen import canvas
from io import BytesIO
import base64 
from fpdf import FPDF
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from Calculation import Calculation

# Define translation dictionaries
translations = {
    "de": {
        "title": "Bemessungssoftware (V1.4.3)",
        "select_language": "Sprache auswählen",
        "variant": "Ausführungsvariante",
        "BT": "Bügeltiefe BT",
        "A": "A (120mm, vorwiegend Querkraft)",
        "geometry": "Geometrie",
        "rib height": "Tragbügelhöhe TBH (cm)",
        "cover down": "Dämmmass unten Du (mm)",
        "cover top": "Dämmmass oben Do (mm)",
        "insulation thickness": "Dämmstoffdicke B (cm)",
        "element lenght": "Elementlänge L",
        "title group3": "Dämmmaterial & Brandschutz",
        "title group4": "Beton",
        "insulation material": "Dämmmaterial",
        "concrete type": "Betonsorte",
        "fire resistance": "Brandschutzklasse",
        "slab thickness": "Minimale Bauteildicke (mm)",
        "title group actions": "Schnittkräfte", 
        "select moment": "Wählen Med (kNm/Element)",
        "select shear": "Wählen Ved (kN/Element)",
        "select axial": "Wählen Ned (kN/Element)",
        "select Horizontal": "Wählen Hed (kN/Element)",
        "run": "Start Bemessung",
        "COMPACT": "Kompakt",
        "wool": "SW",
        "rib number": "Tragbügelanzahl",
        "designation": "Bezeichnung",
        "required N =": "Erforderliche Tragbügelanzahl TBA",
        "actual N =": "Tatsächliche Tragbügelanzahl TBA",
        "error1 =": "Die Anzahl der Rippen darf nicht größer als 10 sein",
        "error2 =": "Die Anzahl der Rippen darf nicht größer als 5 sein",
        "lenght": "länge (cm)",
        "Save Calculation": "Berechnung/Position speichern",
        "Saved Calculation": "Gespeicherte Berechnung/Position",
        "Date": "Datum",
        "Project Name": "Objekt",
        "Component": "Bauteil",
        "Template": "Template.xlsx" ,
        "Order List Data":"Daten der Bestellliste",
        "required N_Vertical =": "für vertikale Lasten : Erforderliche Tragbügelanzahl TBA ",
        "actual N_vertical =": "für vertikale Lasten : Tatsächliche Tragbügelanzahl TBA ",
        "actual N_vertical and Horizontal =": "für vertikale und horizontale Lasten : Tatsächliche Tragbügelanzahl TBA ",
        "Horizontal forces analysis" : "Analyse der Horizontalkräfte ist enthalten"
    },
    "fr": {
        "title": "Logiciel de calcul <br>(V1.4.3)",
        "select_language": "Choisir la langue",
        "variant": "Variante d'exécution",
        "BT": "Profondeur de repassage",
        "A": "A (120mm, principalement force transversale)",
        "geometry": "Géométrie",
        "rib height": "Hauteur de l'étrier de support TBH (cm)",
        "cover down": "Cote d'isolation inférieure (mm)",
        "cover top": "Cote d'isolation supérieure (mm)",
        "insulation thickness": "Épaisseur d'isolant (cm)",
        "element lenght": "Longueur de l'élément L",
        "title group3": "Matériau d'isolation & Résistance au feu",
        "title group4": "Béton",
        "insulation material": "Matériau d'isolation",
        "concrete type": "Sorte de béton",
        "fire resistance": "Classe de réaction au feu",
        "slab thickness": "Épaisseur minimale de l'élément (mm)",
        "title group actions": "Efforts de coupe",
        "select moment": "Sélectionner Med (kNm/Élément)",
        "select shear": "Sélectionner Ved (kN/Élément)",
        "select axial": "Sélectionner Ned (kN/Élément)",
        "select Horizontal": "Sélectionner Hed (kN/Element)",
        "run": "Début du calcul",
        "COMPACT": "Compact",
        "wool": "MW",
        "rib number": "Nombre d'étriers de support",
        "designation": "désignation",
        "required N =": "Requis Nombre d'étriers de support TBA",
        "actual N =": "Nombre réel d'étriers de support TBA",
        "error1 =": "Le nombre de nervures ne peut pas être supérieur à 10",
        "error2 =": "Le nombre de nervures ne peut pas être supérieur à 5",
        "lenght": "Longueur (cm)",
        "Save Calculation": "Sauvegarder le calcul/la position",
        "Saved Calculation": "Calcul/Position sauvegardé",
        "Date": "Date",
        "Project Name": "Objet",
        "Component": "Partie",
        "Template": "TemplateFr.xlsx",
        "Order List Data":"Données de la liste de commande",
        "required N_Vertical =": "Pour charges verticales : Requis Nombre d'étriers de support TBA ",
        "actual N_vertical =": "Pour charges verticales : Nombre réel d'étriers de support TBA ",
        "actual N_vertical and Horizontal =": "Pour les charges verticales et horizontales : Nombre réel d'étriers de support TBA ",
        "Horizontal forces analysis" : "L'analyse des forces horizontales est incluse"
    }
}

def fill_template(template_path, calc_results, Project, Date):
    # Load the excel template
    wb = load_workbook(template_path)
    sheet = wb.active

    # Add Project,component and Date
    sheet['B7'] = f'{Project}'
    sheet['B9'] = f'{Component}'
    sheet['H17'] = f'{Date}'

    # Data starting from row 24 
    row = 24
    for result in calc_results:
        sheet.cell(row, 1).value = result['Pos']
        sheet.cell(row, 3).value = result['STK']
        sheet.cell(row, 4).value = result['FIRIKA']
        sheet.cell(row, 8).value = result['L']
        row += 1
    
    # Save the filled template to a new file
    filled_path = "/tmp/Firika_Report_filled.xlsx"
    wb.save(filled_path)
    
    return filled_path

def get_binary_file_downloader_html(bin_file, file_label='File', button_label='Download'):
    with open(bin_file, 'rb') as f:
        data = f.read()
    bin_str = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{file_label}">{button_label}</a>'
    return href

############ Set index for defualt langauge in URL ############
# Get query parameters from the URL
query_params = st.query_params

# Adjusted handling of 'index' parameter
index = int(query_params.get('index', '0'))

# Adjusted handling of 'lang' parameter and default to 'de' if not found or invalid
lang = query_params.get('lang', 'de')
if lang not in ["de", "fr"]:
    lang = 'de'  # default to 'de' if an invalid value is found

# Define the options for the selectbox
options = ["de", "fr"]

# Determine the index based on the lang parameter
lang_index = options.index(lang)

# Set up the language selection
language = st.sidebar.selectbox(
    label="Sprache auswählen / Choisir la langue",
    options=options,
    index=lang_index,
    format_func=lambda lang: "Deutsch" if lang == "de" else "Français"
)

# Retrieve the appropriate translations
trans = translations[language]

# Injecting custom CSS to style the app
st.markdown(
    """
    <style>
    .title {
        color: red; /* Set the color to red */
        font-size: 24px;
        font-weight: bold;
    }
    .header {
        color: red; /* Set the color to red */
        font-size: 20px;
        font-weight: bold;
        border-bottom: 2px solid #FF5733;
        padding-bottom: 10px;
    }
    .error {
        color: red;
        font-weight: bold;
    }
    .success {
        color: green;
        font-weight: bold;
    }
    .container {
        padding: 10px;
        margin: 10px;
        border: 1px solid #ddd;
        border-radius: 10px;
        background-color: #f9f9f9;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Adding the title relative to the sidebar
st.sidebar.markdown(f'<div class="title"><span class="bold-text">FIRIKA</span><br>{trans["title"]}</div>', unsafe_allow_html=True)

Date = st.sidebar.date_input(trans['Date'])
Project = st.sidebar.text_input(trans['Project Name'])
Component = st.sidebar.text_input(trans['Component'])

################################# App Input ###################################

# Define layout of the page/screen with two columns
col1, col2 = st.columns(2)

# First container in the first column
with col1:
    with st.container():
        st.markdown(f"<h4 class='header'>{trans['variant']}</h4>", unsafe_allow_html=True)
        st.image('1.png')
        F_options = ["C", "O", "G", "Q", "Z"]
        F = st.selectbox(trans["variant"], F_options)

        if F in ["O", "Q"]:
            F1_options = [trans["A"], "B (170 mm)", "C (220 mm)"]
            F1 = st.selectbox(trans["BT"], F1_options)
            if F == "O":
                F = f"O{F1[0]}"
            elif F == "Q":
                F = f"Q{F1[0]}"
                
with col2:
    with st.container():
        st.markdown(f"<h4 class='header'>{trans['geometry']}</h4>", unsafe_allow_html=True)
        image_path = '1ab_de.png' if language == "de" else '1ab_fr.png'
        st.image(image_path)
        H_R_options = [11, 13, 15, 17, 19]
        H_R = st.selectbox(trans['rib height'], H_R_options)
        D_o = st.number_input(trans['cover top'], min_value=25, step=5)
        D_u = st.number_input(trans['cover down'], min_value=25, step=5)
        B = st.selectbox(trans['insulation thickness'], [8, 12])
        L = st.selectbox(trans['element lenght'], ["100 cm", "50 cm", trans['COMPACT']])

col3, col4 = st.columns(2)

with col3:
    with st.container():
        st.markdown(f"<h4 class='header'>{trans['title group3']}</h4>", unsafe_allow_html=True)
        DM_options = ["EPS", trans['wool'], "XPS"]
        DM = st.selectbox(trans['insulation material'], DM_options)
        BS_options = ["R0", "REI60", "REI90"] if DM in ["EPS", "XPS"] else ["REI120"]
        BS = st.selectbox(trans['fire resistance'], BS_options)

with col4:
    with st.container():
        st.markdown(f"<h4 class='header'>{trans['title group4']}</h4>", unsafe_allow_html=True)
        Fck_options = ["C25/30", "C30/37"]
        Fck = st.selectbox(trans['concrete type'], Fck_options)
        Fck = 25 if Fck == "C25/30" else 30
        H_slab = st.number_input(trans['slab thickness'], step=10.0, format="%.1f", value=160.0)

with st.container():
    st.markdown(f"<h4 class='header'>{trans['title group actions']}</h4>", unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        M = st.number_input(trans['select moment'], step=1.0, format="%.1f", value=0.0) * 10**6
    with col2:
        V = st.number_input(trans['select shear'], step=1.0, format="%.1f", value=0.0) * 10**3
    with col3:
        N = st.number_input(trans['select axial'], step=1.0, format="%.1f", value=0.0)
    with col4:
        H = st.number_input(trans['select Horizontal'], step=1.0, format="%.1f", value=0.0)
     
with st.container():
    st.markdown(f"<h4 class='header'>{trans['Order List Data']}</h4>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        Pos = st.text_input('Pos.')
    with col2:
        STK = st.number_input('STK', step=1.0, format="%.1f", value=0.0)

# Ensure `calc_results` persists across sessions
if 'calc_results' not in st.session_state:
    st.session_state.calc_results = []



if st.button((trans['run'])):
    if L == "100 cm":
        L = "100"
    if L == "50 cm":
        L = "50" 

    
    n_vertical_loads_n_Total , n_vertical_loads , v_h_Rd_Verical_loads , n_vert_and_Horz_loads , v_h_Rd ,Presence_Horz_Load = Calculation(H_R,B,Fck,H_slab,M,V,N,H)

    Error1 = False  # Initialize Error to False
    Error2 = False  # Initialize Error to False
    Error3 = False  # Initialize Error to False
    Error4 = False  # Initialize Error to False
    
    if L == "100" or L == (trans['COMPACT']):
        if n_vertical_loads > 10:  
            n_vertical_loads = 10
            Error1 = True
            
        if L == (trans['COMPACT']):
            L = n_vertical_loads*10
        
    if L == "50":
        if n_vertical_loads>5:
            n_vertical_loads = 5
            Error2 = True



    #STRINGS
    H_R = str(H_R)
    D_o = str(D_o)
    D_u = str(D_u)
    H_slab = int(H_slab)
    H_slab = str(H_slab)
    B   = str(B)
    L   = str(L)
    n_vertical_loads   = str(n_vertical_loads)


    if Presence_Horz_Load == False:
        if F == "Z":
            FIRIKA = F + "/" + H_slab + "." + B + "/" + L + "/" + DM + "/" + BS
    
        else:
            FIRIKA = F + "/" + n_vertical_loads + "-" + H_R + "/" + D_o + "." + D_u + "." + B + "/" + L + "/" + DM + "/" + BS
       
        st.write('<p style="font-size:24px;"><b> ' + trans["designation"] + ': ' + FIRIKA + '</b></p>', unsafe_allow_html=True)
        
        n_vertical_loads_n_Total = round(n_vertical_loads_n_Total, 2)
        n_vertical_loads_n_Total = str(n_vertical_loads_n_Total)
        
        if F != "Z":
            if Error1 == True:
                st.write(trans['error1 ='])
                
            elif Error2 == True:
                st.write(trans['error2 ='])
            
            else:
                st.write(f"{trans['required N =']} {n_vertical_loads_n_Total}")
                st.write(f"{trans['actual N =']} {n_vertical_loads}")
###############################################################################
    if Presence_Horz_Load == True:
        
        if L == "100" or L == (trans['COMPACT']):

            if n_vert_and_Horz_loads > 10:  
                n_vert_and_Horz_loads = 10
                Error3 = True
                
            if L == (trans['COMPACT']):
                L = n_vert_and_Horz_loads*10
            
        if L == "50":
            if n_vert_and_Horz_loads>5:
                n_vert_and_Horz_loads = 5
                Error4 = True 
        
        n_vert_and_Horz_loads = str(n_vert_and_Horz_loads)

        if F == "Z":
            FIRIKA = F + "/" + H_slab + "." + B + "/" + L + "/" + DM + "/" + BS

        else:
            FIRIKA = F + "/" + n_vert_and_Horz_loads + "-" + H_R + "/" + D_o + "." + D_u + "." + B + "/" + L + "/" + DM + "/" + BS
        
        st.markdown(f'<br><p><strong><em>{trans["Horizontal forces analysis"]}</em></strong></p>',unsafe_allow_html=True)
        st.write('<p style="font-size:24px;"><b> ' + trans["designation"] + ': ' + FIRIKA + '</b></p>', unsafe_allow_html=True)
        
                
        if F != "Z":
            if Error3 == True:
                st.write(trans['error1 ='])
                
            elif Error4 == True:
                st.write(trans['error2 ='])
            
            else:
                n_vertical_loads_n_Total = round(n_vertical_loads_n_Total,2)
                v_h_Rd_Verical_loads = round(v_h_Rd_Verical_loads,2)
                v_h_Rd = round(v_h_Rd,2)
                
                st.write(f"{trans['required N_Vertical =']} = {n_vertical_loads_n_Total}") 
                st.write(f"{trans['actual N_vertical =']} = {n_vertical_loads} , Hrd = {v_h_Rd_Verical_loads} kN") 
                
                st.markdown(f"<p style='color:green;'>{trans['actual N_vertical and Horizontal =']} = {n_vert_and_Horz_loads} , Hrd = {v_h_Rd} kN</p>",unsafe_allow_html=True )
                    
                    
    # Store the latest calculation result in session state
    st.session_state.latest_result = {
        'FIRIKA': FIRIKA,
        'Pos': Pos,
        'STK': STK,
        'L': L,
    }
    
# Display the latest calculation result for user review
if 'latest_result' in st.session_state:
    if st.button(trans["Save Calculation"]):
        st.session_state.calc_results.append(st.session_state.latest_result)
        del st.session_state.latest_result  # Clear latest result after adding
            
        
# Function to display selected results
def display_selected_results(results, index=True): 
    if results:
        st.write(trans["Saved Calculation"])
        headers = ["Pos", "STK", trans['designation'], trans['lenght']]
        data = [[result['Pos'],  f"{result['STK']:.1f}", result['FIRIKA'], result['L']] for result in results]
        df = pd.DataFrame(data, columns=headers)
        
        # Temporarily set display options to hide index
        # Display DataFrame without default row indices
        df.index = np.arange(1, len(df) + 1)
        st.table(df)
        
    else:
        st.write("No results selected.")    

# Show selected results before generating the PDF
if 'calc_results' in st.session_state and st.session_state.calc_results:
    display_selected_results(st.session_state.calc_results)
         
        
if st.button('Download Excel') and 'calc_results' in st.session_state and st.session_state.calc_results:
    template_path = trans['Template']  # Path to your template.xlsx
    filled_excel_path = fill_template(template_path, st.session_state.calc_results, Project, Date)
    st.markdown(get_binary_file_downloader_html(filled_excel_path, "Firika_Report_filled.xlsx", "Download Excel"), unsafe_allow_html=True)      













